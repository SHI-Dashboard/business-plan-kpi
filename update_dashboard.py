#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SHI 대시보드 데이터 생성기 (연간계획 + 실적)
===============================================================
계획 = 기초자료(2025/2026 P&L Business Plan)  ·  실적 = 모니터링 파일

사용법:
    python update_dashboard.py --base "기초자료_EN.xlsx" --monitor "모니터링.xlsx"

산출:
    data.json  (index.html 이 불러오는 데이터). GitHub 저장소에 커밋/푸시하면 웹이 자동 갱신됩니다.

기준:
  · 사업계획(BP) 회계연도 = 2026.01~2026.12
  · KPI 회계연도 = 2025.11, 12  +  2026.01~10
  · 상단 스코어카드 = 연간 총계획 대비 누계실적(연간 진행률)
  · 지표별 대시보드 = 해당월까지 누계계획 대비 실적

주의: 기초자료의 총매출/영업이익 행은 수식이므로, Excel에서 한 번 열어 저장(값 캐시)한
      파일을 사용하세요. (총매출 연간계획이 0으로 나오면 이 경고가 원인입니다.)
"""
import json, argparse
from openpyxl import load_workbook

LABELS = {
    'tot': 'I. Sales', 'op': 'V. Operating Profit',
    'rel': '-. Affiliate MRO Sales', 'nonrel': '-. Non-Affiliate MRO Sales',
    'mach': '-. Machinery Sales (External)', 'std': '-. Purchasing Standardization Sales',
}
KEYS = ['tot', 'op', 'rel', 'nonrel', 'mach', 'std']

def find_rows(ws):
    m = {}
    for r in range(1, 130):
        b = ws.cell(row=r, column=2).value
        if b is None:
            continue
        bs = str(b).strip()
        for k, lab in LABELS.items():
            if bs == lab:
                m[k] = r
    return m

def monthly_plan(ws, r):  # Jan..Dec = cols D(4)..O(15)
    return [ws.cell(row=r, column=c).value or 0 for c in range(4, 16)]

def read_plan(base_path):
    wb = load_workbook(base_path, data_only=True)
    bp26 = wb['2026 P&L Business Plan']; r26 = find_rows(bp26)
    bp25 = wb['2025 P&L Business Plan']; r25 = find_rows(bp25)
    p26 = {k: monthly_plan(bp26, r26[k]) if k in r26 else [0]*12 for k in LABELS}
    p25 = {k: monthly_plan(bp25, r25[k]) if k in r25 else [0]*12 for k in LABELS}
    bp = {k: p26[k][:] for k in KEYS}                                # 2026 Jan-Dec
    kp = {k: [p25[k][10], p25[k][11]] + p26[k][0:10] for k in KEYS}  # 25.11,12 + 26.01-10
    if not sum(v for v in bp['tot'] if isinstance(v, (int, float))):
        print("[경고] 총매출 연간계획이 0입니다. 기초자료를 Excel에서 열어 저장(값 캐시) 후 다시 실행하세요.")
    return bp, kp

def rng(ws, row, c0, c1):
    return [ws.cell(row=row, column=c).value for c in range(c0, c1 + 1)]

def cell(ws, r, c):
    return ws.cell(row=r, column=c).value

def read_actual(mon_path):
    wb = load_workbook(mon_path, data_only=True)
    ws = wb['Sheet1']
    ACT = {'op': (7, 14), 'opm': (6, 13), 'rel': (19, 23),
           'nonrel': (27, 31), 'mach': (35, 39), 'std': (43, 47)}
    m_act = {'tot': {'bp': rng(ws, 6, 26, 37), 'kp': rng(ws, 6, 24, 37)[:12]}}
    for k, (br, kr) in ACT.items():
        m_act[k] = {'bp': rng(ws, br, 7, 18)[:12], 'kp': rng(ws, kr, 5, 18)[:12]}
    cum_act = {
        'tot': {'bp': cell(ws, 6, 22), 'kp': cell(ws, 6, 23)},
        'op':  {'bp': cell(ws, 10, 22), 'kp': cell(ws, 10, 23)},
        'opm': {'bp': cell(ws, 13, 22), 'kp': cell(ws, 13, 23)},
        'rel': {'bp': cell(ws, 19, 19), 'kp': cell(ws, 23, 19)},
        'nonrel': {'bp': cell(ws, 27, 19), 'kp': cell(ws, 31, 19)},
        'mach': {'bp': cell(ws, 35, 19), 'kp': cell(ws, 39, 19)},
        'std': {'bp': cell(ws, 43, 19), 'kp': cell(ws, 47, 19)},
    }
    pdplb = {'sales': cell(ws, 50, 19), 'cogs': cell(ws, 51, 19), 'sga': cell(ws, 52, 19),
             'op': cell(ws, 53, 19), 'opm': cell(ws, 54, 19),
             'm_sales': rng(ws, 50, 7, 18), 'm_cogs': rng(ws, 51, 7, 18),
             'm_sga': rng(ws, 52, 7, 18), 'm_op': rng(ws, 53, 7, 18)}
    months = ['26.01','26.02','26.03','26.04','26.05','26.06','26.07','26.08','26.09','26.10','26.11','26.12']
    last = '—'
    for i, c in enumerate(range(26, 38)):
        if ws.cell(row=6, column=c).value is not None:
            last = months[i]
    return m_act, cum_act, pdplb, last

def s(a, lo=None, hi=None):
    a = a[lo:hi] if (lo is not None or hi is not None) else a
    return sum(v for v in a if isinstance(v, (int, float)))

def ratio(a, b): return (a / b) if b else 0

def main(base_path, mon_path):
    bp, kp = read_plan(base_path)
    m_act, cum_act, pdplb, last = read_actual(mon_path)
    MON, CUM = {}, {}
    for k in KEYS:
        MON[k] = {'bp_plan': bp[k], 'bp_act': m_act[k]['bp'],
                  'kp_plan': kp[k], 'kp_act': m_act[k]['kp']}
        CUM[k] = {'bp_plan_ytd': s(bp[k], 0, 6), 'bp_plan_year': s(bp[k]), 'bp_act': cum_act[k]['bp'],
                  'kp_plan_ytd': s(kp[k], 0, 8), 'kp_plan_year': s(kp[k]), 'kp_act': cum_act[k]['kp']}
    # --- 총매출(합계) = 카테고리 합으로 계산 (원본 합계 행 오류 방지) ---
    # 관계사+비관계사+기계장비+구매표준화의 합. 미래월(전부 None)은 None 유지.
    CATS = ['rel', 'nonrel', 'mach', 'std']
    def _elemsum(field):
        n = len(MON['rel'][field])
        res = []
        for i in range(n):
            nums = [MON[c][field][i] for c in CATS if isinstance(MON[c][field][i], (int, float))]
            res.append(sum(nums) if nums else None)
        return res
    MON['tot']['bp_act'] = _elemsum('bp_act')
    MON['tot']['kp_act'] = _elemsum('kp_act')
    CUM['tot']['bp_act'] = sum(CUM[c]['bp_act'] for c in CATS if isinstance(CUM[c]['bp_act'], (int, float)))
    CUM['tot']['kp_act'] = sum(CUM[c]['kp_act'] for c in CATS if isinstance(CUM[c]['kp_act'], (int, float)))

    def rarr(op, tot): return [ratio(op[i], tot[i]) if isinstance(tot[i], (int, float)) and tot[i] else None for i in range(len(op))]
    def _ropm(field):
        return [ratio(MON['op'][field][i], MON['tot'][field][i]) if isinstance(MON['tot'][field][i], (int, float)) and MON['tot'][field][i] else None for i in range(len(MON['op'][field]))]
    MON['opm'] = {'bp_plan': rarr(bp['op'], bp['tot']), 'bp_act': _ropm('bp_act'),
                  'kp_plan': rarr(kp['op'], kp['tot']), 'kp_act': _ropm('kp_act')}
    CUM['opm'] = {'bp_plan_ytd': ratio(s(bp['op'],0,6), s(bp['tot'],0,6)), 'bp_plan_year': ratio(s(bp['op']), s(bp['tot'])), 'bp_act': ratio(CUM['op']['bp_act'], CUM['tot']['bp_act']),
                  'kp_plan_ytd': ratio(s(kp['op'],0,8), s(kp['tot'],0,8)), 'kp_plan_year': ratio(s(kp['op']), s(kp['tot'])), 'kp_act': ratio(CUM['op']['kp_act'], CUM['tot']['kp_act'])}
    out = {'months': {'bp': ['26.01','26.02','26.03','26.04','26.05','26.06','26.07','26.08','26.09','26.10','26.11','26.12'],
                      'kp': ['25.11','25.12','26.01','26.02','26.03','26.04','26.05','26.06','26.07','26.08','26.09','26.10']},
           'monthly': MON, 'cum': CUM, 'pdplb': pdplb, 'updated': f'{last} 누계 · 연간계획 반영'}
    with open('data.json', 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    print(f"[OK] data.json 생성 완료 · 실적 기준월 {last}")
    print(f"     총매출 연간계획  BP {s(bp['tot']):,.0f} / KPI {s(kp['tot']):,.0f}")

if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--base', required=True, help='기초자료 엑셀 (연간계획)')
    ap.add_argument('--monitor', required=True, help='모니터링 엑셀 (실적)')
    a = ap.parse_args()
    main(a.base, a.monitor)
